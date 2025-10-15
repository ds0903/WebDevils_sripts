from aiogram import Router, F
from aiogram.types import CallbackQuery, FSInputFile
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
import os

import sys
sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from database import Database
from bot.utils import is_authorized, logger
from bot.keyboards_aiogram import stats_menu_markup, back_button_markup

router = Router()
db = Database()

@router.callback_query(F.data == "menu_stats")
async def show_stats_menu(callback: CallbackQuery):
    if not is_authorized(callback.from_user.id):
        await callback.answer("❌ Доступ заборонено!", show_alert=True)
        return
    
    stats = db.get_statistics()
    
    text = "═══════════════════════════════════\n"
    text += "  📊 <b>СТАТИСТИКА ТА ІСТОРІЯ</b>\n"
    text += "═══════════════════════════════════\n\n"
    text += "📈 <b>Загальна статистика:</b>\n\n"
    text += f"Всього коментарів: {stats['total']}\n"
    text += f"✅ Успішних: {stats['success']}\n"
    text += f"⚠️ Невдалих: {stats['failed']}\n"
    text += f"❌ Помилок: {stats['error']}\n"
    
    if stats['total'] > 0:
        success_rate = (stats['success'] / stats['total']) * 100
        text += f"\n📊 Success Rate: {success_rate:.1f}%\n"
    
    text += "\n<b>Завантажити історію в Excel:</b>"
    
    await callback.message.edit_text(text, reply_markup=stats_menu_markup(), parse_mode='HTML')
    await callback.answer()

@router.callback_query(F.data.in_(["stats_20", "stats_50"]))
async def view_history(callback: CallbackQuery):
    if not is_authorized(callback.from_user.id):
        await callback.answer("❌ Доступ заборонено!", show_alert=True)
        return
    
    limit = 20 if callback.data == "stats_20" else 50
    history = db.get_comment_history(limit)
    
    if not history:
        await callback.answer("⚠️ Історії ще немає", show_alert=True)
        return
    
    await callback.answer("📊 Генерую Excel файл...", show_alert=False)
    
    try:
        # Створюємо Excel файл
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Історія {limit}"
        
        # Заголовки
        headers = ['№', 'Дата', 'Час', 'Акаунт', 'Статус', 'Ключ.слово', 'Коментар', 'Посилання на пост', 'Post ID', 'Помилка']
        
        # Стилі заголовків
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Заповнюємо дані
        for idx, item in enumerate(history, 2):
            created_at = item['created_at']
            if isinstance(created_at, str):
                created_at = datetime.fromisoformat(created_at)
            
            date_str = created_at.strftime('%d.%m.%Y')
            time_str = created_at.strftime('%H:%M:%S')
            
            # Статус з emoji
            status_text = {
                'success': '✅ Успішно',
                'failed': '⚠️ Невдало',
                'error': '❌ Помилка'
            }.get(item['status'], '❓')
            
            ws.cell(row=idx, column=1, value=idx-1)
            ws.cell(row=idx, column=2, value=date_str)
            ws.cell(row=idx, column=3, value=time_str)
            ws.cell(row=idx, column=4, value=f"@{item['username']}")
            ws.cell(row=idx, column=5, value=status_text)
            ws.cell(row=idx, column=6, value=item['keyword'] or '')
            ws.cell(row=idx, column=7, value=item['comment_text'] or '')
            ws.cell(row=idx, column=8, value=item['post_link'] or '')
            ws.cell(row=idx, column=9, value=item['post_id'] or '')
            ws.cell(row=idx, column=10, value=item.get('error_message') or '')
            
            # Колір рядка залежно від статусу
            if item['status'] == 'success':
                fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            elif item['status'] == 'failed':
                fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            else:
                fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
            
            for col in range(1, len(headers) + 1):
                ws.cell(row=idx, column=col).fill = fill
        
        # Автоширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Заморожуємо верхній рядок
        ws.freeze_panes = 'A2'
        
        # Зберігаємо файл
        data_dir = Path('data')
        data_dir.mkdir(exist_ok=True)
        
        filename = f"data/history_{limit}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        
        # Відправляємо файл
        file = FSInputFile(filename)
        await callback.message.answer_document(
            file,
            caption=f"📊 <b>Історія коментарів (останні {limit})</b>\n\n"
                   f"Всього записів: {len(history)}\n"
                   f"✅ Успішних: {sum(1 for x in history if x['status'] == 'success')}\n"
                   f"⚠️ Невдалих: {sum(1 for x in history if x['status'] == 'failed')}\n"
                   f"❌ Помилок: {sum(1 for x in history if x['status'] == 'error')}",
            parse_mode='HTML',
            reply_markup=back_button_markup()
        )
        
        # Видаляємо файл після відправки
        try:
            os.remove(filename)
        except:
            pass
        
        logger.info(f"Відправлено Excel файл історії (останні {limit})")
        
    except Exception as e:
        await callback.message.answer(f"❌ Помилка створення файлу: {e}")
        logger.error(f"Помилка створення Excel: {e}")
